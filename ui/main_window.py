"""메인 윈도우 모듈 - 3열 레이아웃"""
import customtkinter as ctk
from tkinter import filedialog
import threading
from core.parser import IPParser
from core.matcher import Matcher
from ui.input_panel import InputPanel
from ui.result_grid import ResultGrid
from ui.title_bar import CustomTitleBar
import pandas as pd
import platform


class MainWindow:
    """메인 윈도우 클래스"""
    
    def __init__(self):
        self.root = ctk.CTk()
        self.root.title("IP Network Matcher")
        
        # 윈도우 커스터마이징
        self.setup_window()
        
        # 데이터 저장
        self.source_data = []
        self.reference_data = []
        
        self.setup_ui()
    
    def setup_window(self):
        """윈도우 설정 및 커스터마이징"""
        # 기본 타이틀 바 제거 (커스텀 타이틀 바 사용)
        self.root.overrideredirect(True)
        
        # 윈도우 크기
        self.root.geometry("1400x700")
        self.root.minsize(1200, 600)
        
        # 투명도 효과 제거 (성능 향상)
        # self.root.attributes("-alpha", 0.97)
        
        # 윈도우 중앙 배치
        self.center_window()
        
        # 배경색 설정 (미니멀 다크 테마)
        self.root.configure(bg="#1a1a1a")
    
    def center_window(self):
        """윈도우를 화면 중앙에 배치"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def setup_ui(self):
        """UI 구성 - 3열 레이아웃"""
        # 커스텀 타이틀 바 (경계 없이)
        self.title_bar = CustomTitleBar(
            self.root,
            self.root,
            on_close=self.on_close
        )
        self.title_bar.pack(fill="x", side="top", padx=0, pady=0)
        
        # 메인 컨테이너 (미니멀 다크 테마)
        main_container = ctk.CTkFrame(
            self.root,
            fg_color=("#1a1a1a", "#1a1a1a"),
            corner_radius=0,
            border_width=0
        )
        main_container.pack(fill="both", expand=True, padx=0, pady=0)
        
        # 내부 컨테이너
        inner_container = ctk.CTkFrame(main_container, fg_color="transparent")
        inner_container.pack(fill="both", expand=True, padx=12, pady=12)
        
        # 컨트롤 버튼 영역 (미니멀 디자인)
        control_frame = ctk.CTkFrame(inner_container, fg_color="transparent")
        control_frame.pack(fill="x", pady=(0, 8))
        
        # 버튼 그룹 (오른쪽 정렬)
        button_group = ctk.CTkFrame(control_frame, fg_color="transparent")
        button_group.pack(side="right")
        
        self.analyze_btn = ctk.CTkButton(
            button_group,
            text="분석",
            command=self.start_analysis,
            font=ctk.CTkFont(size=11, weight="bold"),
            height=30,
            width=70,
            corner_radius=4,
            fg_color=("#333333", "#333333"),
            hover_color=("#444444", "#444444"),
            text_color=("#e0e0e0", "#e0e0e0")
        )
        self._loading_active = False  # 로딩 상태 플래그
        self.analyze_btn.pack(side="left", padx=(0, 4))
        
        self.export_btn = ctk.CTkButton(
            button_group,
            text="저장",
            command=self.export_to_excel,
            font=ctk.CTkFont(size=11),
            height=30,
            width=70,
            corner_radius=4,
            fg_color=("#333333", "#333333"),
            hover_color=("#444444", "#444444"),
            text_color=("#e0e0e0", "#e0e0e0"),
            state="disabled"
        )
        self.export_btn.pack(side="left", padx=(0, 4))
        
        self.reset_btn = ctk.CTkButton(
            button_group,
            text="초기화",
            command=self.reset_all,
            font=ctk.CTkFont(size=11),
            height=30,
            width=60,
            corner_radius=4,
            fg_color=("#333333", "#333333"),
            hover_color=("#444444", "#444444"),
            text_color=("#e0e0e0", "#e0e0e0")
        )
        self.reset_btn.pack(side="left")
        
        # 3열 레이아웃
        columns_frame = ctk.CTkFrame(inner_container, fg_color="transparent")
        columns_frame.pack(fill="both", expand=True)
        
        # 좌측: Source 패널
        self.source_panel = InputPanel(
            columns_frame,
            "Source",
            is_reference=False,
            on_data_change=self.on_data_change
        )
        self.source_panel.pack(side="left", fill="both", expand=True, padx=(0, 4))
        
        # 중앙: Reference 패널
        self.reference_panel = InputPanel(
            columns_frame,
            "Reference",
            is_reference=True,
            on_data_change=self.on_data_change
        )
        self.reference_panel.pack(side="left", fill="both", expand=True, padx=(4, 4))
        
        # 우측: 결과 패널
        self.result_grid = ResultGrid(columns_frame)
        self.result_grid.pack(side="left", fill="both", expand=True, padx=(4, 0))
        
        # 하단 상태 바
        status_frame = ctk.CTkFrame(inner_container, fg_color="transparent")
        status_frame.pack(fill="x", pady=(8, 0))
        
        # 로딩 인디케이터 (초기에는 숨김)
        self.loading_label = ctk.CTkLabel(
            status_frame,
            text="●",
            font=ctk.CTkFont(size=8),
            text_color=("#4ade80", "#4ade80")
        )
        self.loading_label.pack(side="left", padx=(0, 6))
        self.loading_label.pack_forget()  # 초기에는 숨김
        
        self.progress_label = ctk.CTkLabel(
            status_frame,
            text="준비됨",
            font=ctk.CTkFont(size=10),
            text_color=("#888888", "#888888")
        )
        self.progress_label.pack(side="left")
    
    def on_data_change(self):
        """데이터 변경 이벤트 핸들러"""
        # 패널이 아직 생성되지 않았을 수 있으므로 확인
        if not hasattr(self, 'source_panel') or not hasattr(self, 'reference_panel'):
            return
        
        source_text = self.source_panel.get_text_content().strip()
        reference_text = self.reference_panel.get_text_content().strip()
        
        if source_text and reference_text:
            self.analyze_btn.configure(state="normal")
        else:
            self.analyze_btn.configure(state="disabled")
    
    def start_analysis(self):
        """분석 시작"""
        self.analyze_btn.configure(state="disabled", text="분석 중...")
        self.progress_label.configure(text="분석 중...", text_color=("#888888", "#888888"))
        
        # 로딩 인디케이터 표시
        self._loading_active = True
        self.loading_label.pack(side="left", padx=(0, 6))
        self._animate_loading()
        
        # 별도 스레드에서 실행 (UI 멈춤 방지)
        thread = threading.Thread(target=self.perform_analysis)
        thread.daemon = True
        thread.start()
    
    def _animate_loading(self):
        """로딩 애니메이션"""
        if self._loading_active:
            # 간단한 깜빡임 효과
            current_color = self.loading_label.cget("text_color")[0]
            new_color = "#1a1a1a" if current_color == "#4ade80" else "#4ade80"
            self.loading_label.configure(text_color=(new_color, new_color))
            self.root.after(500, self._animate_loading)
    
    def perform_analysis(self):
        """실제 분석 수행"""
        try:
            # Source 데이터 파싱
            source_text = self.source_panel.get_text_content()
            self.source_data = IPParser.parse_text_input(source_text)
            
            # Reference 데이터 파싱
            reference_text = self.reference_panel.get_text_content()
            parsed_refs = IPParser.parse_text_input(reference_text)
            self.reference_data = []
            for ref in parsed_refs:
                self.reference_data.append({
                    'original': ref['original'],
                    'parsed': ref['parsed'],
                    'type': ref['type']
                })
            
            # 매칭 수행
            results = Matcher.match(self.source_data, self.reference_data)
            
            # UI 업데이트 (메인 스레드에서 실행)
            self.root.after(0, self.update_results, results)
            
        except Exception as e:
            error_msg = f"분석 오류: {str(e)}"
            print(error_msg)
            self.root.after(0, self.show_error, error_msg)
    
    def update_results(self, results):
        """결과 업데이트"""
        # 로딩 인디케이터 숨김
        self._loading_active = False
        self.loading_label.pack_forget()
        
        self.result_grid.display_results(results)
        matched_count = sum(1 for r in results if r.get('matched_ips', '').strip())
        total_count = len(results)
        self.progress_label.configure(
            text=f"완료: {matched_count}/{total_count}",
            text_color=("#888888", "#888888")
        )
        self.analyze_btn.configure(state="normal", text="분석")
        self.export_btn.configure(state="normal")
    
    def show_error(self, error_msg):
        """오류 메시지 표시"""
        # 로딩 인디케이터 숨김
        self._loading_active = False
        self.loading_label.pack_forget()
        
        self.progress_label.configure(
            text=f"오류: {error_msg[:50]}",
            text_color=("#f87171", "#f87171")
        )
        self.analyze_btn.configure(state="normal", text="분석")
    
    def export_to_excel(self):
        """엑셀로 내보내기 (깔끔한 서식)"""
        results = self.result_grid.get_results_data()
        if not results:
            return
        
        file_path = filedialog.asksaveasfilename(
            title="엑셀 파일로 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # 워크북 생성
                wb = Workbook()
                ws = wb.active
                ws.title = "매칭 결과"
                
                # 헤더 설정
                headers = ['대상 IP', '매칭된 IP']
                ws.append(headers)
                
                # 헤더 스타일 설정
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                
                # 데이터 입력
                for result in results:
                    source = result.get('source', '')
                    matched_ips = result.get('matched_ips', '')
                    ws.append([source, matched_ips])
                
                # 데이터 셀 스타일 설정
                data_font = Font(size=10)
                data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
                    for cell in row:
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = border
                
                # 컬럼 너비 자동 조정
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 50
                
                # 행 높이 설정
                ws.row_dimensions[1].height = 25  # 헤더 높이
                for row_num in range(2, ws.max_row + 1):
                    ws.row_dimensions[row_num].height = 20
                
                # 저장
                wb.save(file_path)
                
                file_name = file_path.split('/')[-1]
                self.progress_label.configure(
                    text=f"저장됨: {file_name}",
                    text_color=("#888888", "#888888")
                )
            except Exception as e:
                error_msg = f"저장 오류: {str(e)}"
                print(error_msg)
                self.progress_label.configure(
                    text=f"오류: {error_msg[:50]}",
                    text_color=("#f87171", "#f87171")
                )
    
    def reset_all(self):
        """전체 초기화"""
        self.source_panel.clear_data()
        self.reference_panel.clear_data()
        self.result_grid.display_results([])
        self.progress_label.configure(text="준비됨", text_color=("#888888", "#888888"))
        self.export_btn.configure(state="disabled")
        self.source_data = []
        self.reference_data = []
    
    def on_close(self):
        """윈도우 닫기"""
        self.root.destroy()
    
    def run(self):
        """애플리케이션 실행"""
        self.root.mainloop()
